from openpyxl import load_workbook, Workbook
from openpyxl.utils.cell import column_index_from_string
# from openpyxl.utils import cell
import xlwings as xw
import time


def apply_Formula_to_Run(exfile):
    try:
        
        wb1 = load_workbook(exfile)
        ws1 = wb1.worksheets[0]
        formula_list = [
            "=$F$4", "=$J$4", "=$N$4", "=$R$4", "=$V$4", "=$Z$4", "=$AD$4", "=$AH$4",
            "=$AL$4", "=$AP$4", "=$AT$4", "=$H$3", "=$L$3", "=$P$3", "=$T$3", "=$X$3",
            "=$AB$3", "=$AF$3", "=$AJ$3", "=$AN$3", "=$AR$3", "=$AV$3", "=$I$3", "=$M$3",
            "=$Q$3", "=$U$3", "=$Y$3", "=$AC$3", "=$AG$3", "=$AK$3", "=$AO$3", "=$AS$3",
            "=$AW$3", "=$AZ$3", "=$AZ$4", "=$BA$4"
        ]
        row = 7
        for i in formula_list:
            ws1.cell(row=row, column=1).value = i
            row +=1
        wb1.save(exfile)    
        wb1.close()

    except Exception as e:
        print("An error occurred while updating the formula:", e)
        error_message  = str(e)
        start_index = error_message.find("'") + 1
        end_index = error_message.rfind(".xlsx'") + len(".xlsx")
        extracted_filename = error_message[start_index:end_index]
        print(extracted_filename)
        # closing workbooks if opened with same names

