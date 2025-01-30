from openpyxl import load_workbook, Workbook
from openpyxl.utils.cell import column_index_from_string
# from openpyxl.utils import cell
import xlwings as xw
import time




def copy_from_awb_file(exawbfile,extempfile):
    try:
        print("Loading workbook")
        wb1 = load_workbook(exawbfile)
        ws1 = wb1.worksheets[0]
        wbtemp = load_workbook(extempfile)
        wstemp = wbtemp.worksheets[0]
        wstemp2 = wbtemp.worksheets[1]
        wstemp3 = wbtemp.worksheets[2]
        output_row = 208
        # column 10,11,12
        ranges = [
                (3, 10, 4),
                (13, 20, 5),
                (23, 30, 6),
                (33, 40, 7),
                (43, 50, 8),
                (53, 60, 9)
            ]
        value1  = []
        print("Copying data")
        for range_start, range_end, source_column in ranges:
            for row in range(range_start, range_end + 1):
                value1.append(ws1.cell(row=row, column=10).value)  # Store values for the range  
            for idx, val in enumerate(value1):
                row = output_row + idx
                wstemp.cell(row=row, column=source_column).value = val
        
            value1.clear()

        wbtemp.save(extempfile)
        
        ranges1 = [
                (3, 10, 4),
                (13, 20, 5),
                (23, 30, 6),
                (33, 40, 7),
                (43, 50, 8),
                (53, 60, 9)
            ]
        output_row1 = 216

        value2 = []

        for range_start, range_end, source_column in ranges1:
            for row in range(range_start, range_end + 1):
                value2.append(ws1.cell(row=row, column=11).value)
                  
            for idx, val in enumerate(value2):
                row = output_row1 + idx
                wstemp.cell(row=row, column=source_column).value = val
        
            value2.clear()
        wbtemp.save(extempfile)


        ranges2 = [
                (3, 10, 4),
                (13, 20, 5),
                (23, 30, 6),
                (33, 40, 7),
                (43, 50, 8),
                (53, 60, 9)
            ]
        value3 = []
        output_row2 = 224
        for range_start, range_end, source_column in ranges2:
            for row in range(range_start, range_end + 1):
                value3.append(ws1.cell(row=row, column=12).value)  # Store values for the range  
            for idx, val in enumerate(value3):
                row = output_row2 + idx
                wstemp.cell(row=row, column=source_column).value = val
        
            value3.clear()

        wbtemp.save(extempfile)
        print("Copying alexis awb data completed")

        output_row = 208
        # column 10,11,12
        ranges = [
                (63, 70, 4),
                (73, 80, 5),
                (83, 90, 6),
                (93, 100, 7),
                (103, 110, 8),
                (113, 120, 9)
            ]
        value1  = []
        print("Copying noface awb data")
        for range_start, range_end, source_column in ranges:
            for row in range(range_start, range_end + 1):
                value1.append(ws1.cell(row=row, column=10).value)  # Store values for the range  
            for idx, val in enumerate(value1):
                row = output_row + idx
                wstemp2.cell(row=row, column=source_column).value = val
        
            value1.clear()

        wbtemp.save(extempfile)
        
        ranges1 = [
                (63, 70, 4),
                (73, 80, 5),
                (83, 90, 6),
                (93, 100, 7),
                (103, 110, 8),
                (113, 120, 9)
            ]
        output_row1 = 216

        value2 = []

        for range_start, range_end, source_column in ranges1:
            for row in range(range_start, range_end + 1):
                value2.append(ws1.cell(row=row, column=11).value)
                  
            for idx, val in enumerate(value2):
                row = output_row1 + idx
                wstemp2.cell(row=row, column=source_column).value = val
        
            value2.clear()
        wbtemp.save(extempfile)


        ranges2 = [
                (63, 70, 4),
                (73, 80, 5),
                (83, 90, 6),
                (93, 100, 7),
                (103, 110, 8),
                (113, 120, 9)
            ]
        value3 = []
        output_row2 = 224
        for range_start, range_end, source_column in ranges2:
            for row in range(range_start, range_end + 1):
                value3.append(ws1.cell(row=row, column=12).value)  # Store values for the range  
            for idx, val in enumerate(value3):
                row = output_row2 + idx
                wstemp2.cell(row=row, column=source_column).value = val
        
            value3.clear()

        wbtemp.save(extempfile)
        print("Copying noface awb data completed")

        output_row = 208
        # column 10,11,12
        ranges = [
                (123, 130, 4),
                (133, 140, 5),
                (143, 150, 6),
                (153, 160, 7),
                (163, 170, 8),
                (173, 180, 9)
            ]
        value1  = []
        print("Copying noface awb data")
        for range_start, range_end, source_column in ranges:
            for row in range(range_start, range_end + 1):
                value1.append(ws1.cell(row=row, column=10).value)  # Store values for the range  
            for idx, val in enumerate(value1):
                row = output_row + idx
                wstemp3.cell(row=row, column=source_column).value = val
        
            value1.clear()

        wbtemp.save(extempfile)
        
        ranges1 = [
                (123, 130, 4),
                (133, 140, 5),
                (143, 150, 6),
                (153, 160, 7),
                (163, 170, 8),
                (173, 180, 9)
            ]
        output_row1 = 216

        value2 = []

        for range_start, range_end, source_column in ranges1:
            for row in range(range_start, range_end + 1):
                value2.append(ws1.cell(row=row, column=11).value)
                  
            for idx, val in enumerate(value2):
                row = output_row1 + idx
                wstemp3.cell(row=row, column=source_column).value = val
        
            value2.clear()
        wbtemp.save(extempfile)


        ranges2 = [
                (123, 130, 4),
                (133, 140, 5),
                (143, 150, 6),
                (153, 160, 7),
                (163, 170, 8),
                (173, 180, 9)
            ]
        value3 = []
        output_row2 = 224
        for range_start, range_end, source_column in ranges2:
            for row in range(range_start, range_end + 1):
                value3.append(ws1.cell(row=row, column=12).value)  # Store values for the range  
            for idx, val in enumerate(value3):
                row = output_row2 + idx
                wstemp3.cell(row=row, column=source_column).value = val
        
            value3.clear()

        wbtemp.save(extempfile)
        print("Copying richard awb data completed")

    except Exception as e:
        print("An error occurred:", e)
        error_message  = str(e)
        start_index = error_message.find("'") + 1
        end_index = error_message.rfind(".xlsx'") + len(".xlsx")

        extracted_filename = error_message[start_index:end_index]
        print(extracted_filename)

