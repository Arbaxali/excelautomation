from openpyxl import load_workbook, Workbook
from openpyxl.utils.cell import column_index_from_string
# from openpyxl.utils import cell
import xlwings as xw
import time


def copy_from_ae_file(exaefile, extempfile):
    try:
        print("loading ae file")
        ranges = [
                (13, 18, 4),
                (21, 26, 5),
                (29, 34, 6),
                (37, 42, 7),
                (45, 50, 8),
                (53, 58, 9)
            ]
        wb1 = load_workbook(exaefile)
        ws1 = wb1.worksheets[0]
        wbtemp = load_workbook(extempfile)
        wstemp = wbtemp.worksheets[0]
        output_row = 190  
        value1 = [] 


        print("Copying ae_seqtime data")
        for range_start, range_end, source_column in ranges:
            for row in range(range_start, range_end + 1):
                value1.append(ws1.cell(row=row, column=7).value)  # Store values for the range  
            for idx, val in enumerate(value1):
                row = output_row + idx
                wstemp.cell(row=row, column=source_column).value = val
        
            value1.clear()

        wbtemp.save(extempfile)

        ranges1 = [
                (13, 18, 4),
                (21, 26, 5),
                (29, 34, 6),
                (37, 42, 7),
                (45, 50, 8),
                (53, 58, 9)
            ]
        
        value2 = [] 
        output_row1 = 196  
        print("Copying ae_seqTarget data")
        for range_start, range_end, source_column in ranges1:
            for row in range(range_start, range_end + 1):
                value2.append(ws1.cell(row=row, column=8).value) 
            for idx, val in enumerate(value2):
                row = output_row1 + idx
                wstemp.cell(row=row, column=source_column).value = val
        
            value2.clear()

        wbtemp.save(extempfile)

        ranges2 = [
                (13, 18, 4),
                (21, 26, 5),
                (29, 34, 6),
                (37, 42, 7),
                (45, 50, 8),
                (53, 58, 9)
            ]
        
        value3 = [] 
        output_row2 = 202 
        print("Copying ae_seqStability data")
        for range_start, range_end, source_column in ranges2:
            for row in range(range_start, range_end + 1):
                value3.append(ws1.cell(row=row, column=9).value) 
            for idx, val in enumerate(value3):
                row = output_row2 + idx
                wstemp.cell(row=row, column=source_column).value = val
        
            value3.clear()

        wbtemp.save(extempfile)
        print("saving temp file")
        wbtemp.close()
        # no face data
        ranges = [
                (61, 66, 4),
                (69, 74, 5),
                (77, 82, 6),
                (85, 90, 7),
                (93, 98, 8),
                (101, 106, 9)
            ]
        
        wbtemp = load_workbook(extempfile)
        wstemp2 = wbtemp.worksheets[1]
        output_row = 190  
        value1 = [] 


        print("Copying no face ae_seqtime data")
        for range_start, range_end, source_column in ranges:
            for row in range(range_start, range_end + 1):
                value1.append(ws1.cell(row=row, column=7).value)  # Store values for the range  
            for idx, val in enumerate(value1):
                row = output_row + idx
                wstemp2.cell(row=row, column=source_column).value = val
        
            value1.clear()

        wbtemp.save(extempfile)

        ranges1 = [
                (61, 66, 4),
                (69, 74, 5),
                (77, 82, 6),
                (85, 90, 7),
                (93, 98, 8),
                (101, 106, 9)
            ]
        
        value2 = [] 
        output_row1 = 196  
        print("Copying noface ae_seqTarget data")
        for range_start, range_end, source_column in ranges1:
            for row in range(range_start, range_end + 1):
                value2.append(ws1.cell(row=row, column=8).value) 
            for idx, val in enumerate(value2):
                row = output_row1 + idx
                wstemp2.cell(row=row, column=source_column).value = val
        
            value2.clear()

        wbtemp.save(extempfile)

        ranges2 = [
                (61, 66, 4),
                (69, 74, 5),
                (77, 82, 6),
                (85, 90, 7),
                (93, 98, 8),
                (101, 106, 9)
            ]
        
        value3 = [] 
        output_row2 = 202 
        print("Copying no face ae_seqStability data")
        for range_start, range_end, source_column in ranges2:
            for row in range(range_start, range_end + 1):
                value3.append(ws1.cell(row=row, column=9).value) 
            for idx, val in enumerate(value3):
                row = output_row2 + idx
                wstemp2.cell(row=row, column=source_column).value = val
        
            value3.clear()

        wbtemp.save(extempfile)
        print("saving no face temp file")
       
       # richard data
        ranges = [
                (109, 114, 4),
                (117, 122, 5),
                (125, 130, 6),
                (133, 138, 7),
                (141, 146, 8),
                (149, 154, 9)
            ]
        
        wbtemp = load_workbook(extempfile)
        wstemp3 = wbtemp.worksheets[2]
        output_row = 190  
        value1 = [] 


        print("Copying Richard ae_seqtime data")
        for range_start, range_end, source_column in ranges:
            for row in range(range_start, range_end + 1):
                value1.append(ws1.cell(row=row, column=7).value)  # Store values for the range  
            for idx, val in enumerate(value1):
                row = output_row + idx
                wstemp3.cell(row=row, column=source_column).value = val
        
            value1.clear()

        wbtemp.save(extempfile)

        ranges1 = [
                (109, 114, 4),
                (117, 122, 5),
                (125, 130, 6),
                (133, 138, 7),
                (141, 146, 8),
                (149, 154, 9)
            ]
        
        value2 = [] 
        output_row1 = 196  
        print("Copying richard ae_seqTarget data")
        for range_start, range_end, source_column in ranges1:
            for row in range(range_start, range_end + 1):
                value2.append(ws1.cell(row=row, column=8).value) 
            for idx, val in enumerate(value2):
                row = output_row1 + idx
                wstemp3.cell(row=row, column=source_column).value = val
        
            value2.clear()

        wbtemp.save(extempfile)

        ranges2 = [
                (109, 114, 4),
                (117, 122, 5),
                (125, 130, 6),
                (133, 138, 7),
                (141, 146, 8),
                (149, 154, 9)
            ]
        
        value3 = [] 
        output_row2 = 202 
        print("Copying richard ae_seqStability data")
        for range_start, range_end, source_column in ranges2:
            for row in range(range_start, range_end + 1):
                value3.append(ws1.cell(row=row, column=9).value) 
            for idx, val in enumerate(value3):
                row = output_row2 + idx
                wstemp3.cell(row=row, column=source_column).value = val
        
            value3.clear()

        wbtemp.save(extempfile)
        print("saving richard data temp file")

        wb1.close()
        wbtemp.close()
        
    except Exception as e:
        print("An error occurred:", e)
        error_message  = str(e)
        start_index = error_message.find("'") + 1
        end_index = error_message.rfind(".xlsx'") + len(".xlsx")

        extracted_filename = error_message[start_index:end_index]
        print(extracted_filename)
