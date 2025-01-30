from openpyxl import load_workbook, Workbook
from openpyxl.utils.cell import column_index_from_string
# from openpyxl.utils import cell
import xlwings as xw
import time



def read_AppliedFormulaDatat_to_alexis_Add_to_Temp(exfileReadFormulaData,excel_file_temp):
    
    try:
        
        sheetColumnFlag = ''
        sheetRowFlag = 0                                                    # d=4,e =5,f=6,g=7,h=8,i=9
        FlagFileName = exfileReadFormulaData
        if "Alexis_Challenge_A_20lux_720p_Run_1" in FlagFileName:
                sheetColumnFlag = 7
                sheetRowFlag = 6
        elif "Alexis_Challenge_A_20lux_720p_Run_2" in FlagFileName: 
                sheetColumnFlag = 8
                sheetRowFlag = 6
        elif "Alexis_Challenge_A_20lux_720p_Run_3" in FlagFileName:
                sheetColumnFlag = 9
                sheetRowFlag = 6         
        elif "Alexis_Challenge_A_20lux_1080p_Run_1" in FlagFileName:
                sheetColumnFlag = 4
                sheetRowFlag = 6  
        elif "Alexis_Challenge_A_20lux_1080p_Run_2" in FlagFileName:
                sheetColumnFlag = 5
                sheetRowFlag = 6 
        elif "Alexis_Challenge_A_20lux_1080p_Run_3" in FlagFileName:
                sheetColumnFlag = 6
                sheetRowFlag = 6 
        elif "Alexis_Challenge_CW_80lux_720p_Run_1" in FlagFileName:
                sheetColumnFlag = 7
                sheetRowFlag = 42
        elif "Alexis_Challenge_CW_80lux_720p_Run_2" in FlagFileName:
                sheetColumnFlag = 8
                sheetRowFlag = 42 
        elif "Alexis_Challenge_CW_80lux_720p_Run_3" in FlagFileName:
                sheetColumnFlag = 9
                sheetRowFlag = 42  
        elif "Alexis_Challenge_CW_80lux_1080p_Run_1" in FlagFileName:        
                sheetColumnFlag = 4
                sheetRowFlag = 42 
        elif "Alexis_Challenge_CW_80lux_1080p_Run_2" in FlagFileName: 
                sheetColumnFlag = 5
                sheetRowFlag = 42 
                print('Sheet E42') 
        elif "Alexis_Challenge_CW_80lux_1080p_Run_3" in FlagFileName: 
                sheetColumnFlag = 6
                sheetRowFlag = 42 
        elif "Alexis_Challenge_D65_250lux_720p_Run_1" in FlagFileName:                  # d=4,e =5,f=6,g=7,h=8,i=9
                sheetColumnFlag = 7
                sheetRowFlag = 78 
        elif "Alexis_Challenge_D65_250lux_720p_Run_2" in FlagFileName: 
                sheetColumnFlag = 8
                sheetRowFlag = 78
        elif "Alexis_Challenge_D65_250lux_720p_Run_3" in FlagFileName: 
                sheetColumnFlag = 9
                sheetRowFlag = 78   
        elif "Alexis_Challenge_D65_250lux_1080p_Run_1" in FlagFileName:      
                sheetColumnFlag = 4
                sheetRowFlag = 78 
        elif "Alexis_Challenge_D65_250lux_1080p_Run_2" in FlagFileName:  
                sheetColumnFlag = 5
                sheetRowFlag = 78 
        elif "Alexis_Challenge_D65_250lux_1080p_Run_3" in FlagFileName: 
                sheetColumnFlag = 6
                sheetRowFlag = 78


        # wxb1 = xw.Book(exfileReadFormulaData)
        # wxb1s = wxb1.sheets['Sheet1']
        wxb1 = load_workbook(exfileReadFormulaData)
        wxb1s = wxb1.worksheets[0]

        # list is declared
        AlexisExcelData = [] 
        for i, row in enumerate(range(7, 43)):
            AlexisExcelData.append(wxb1s.cell(row=row, column=1).value)
        wxb1.close()


        wxbtemp = load_workbook(excel_file_temp)
        wxb1s = wxbtemp.worksheets[0]
        n = 0
        for Excelvalue in AlexisExcelData:           
            wxb1s.cell(row = sheetColumnFlag,  column= sheetRowFlag).value = Excelvalue
            n += 1


        wxbtemp.save(excel_file_temp)
        wxbtemp.close()


    except Exception as e:
        print("An error occurred while reading and updaing temp excel:", e)
        error_message  = str(e)
        start_index = error_message.find("'") + 1
        end_index = error_message.rfind(".xlsx'") + len(".xlsx")
        extracted_filename = error_message[start_index:end_index]
        print(extracted_filename)
        # closing workbooks if opened with same names


