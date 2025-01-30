from openpyxl import load_workbook, Workbook
from openpyxl.utils.cell import column_index_from_string
# from openpyxl.utils import cell
import xlwings as xw
import time


def read_AppliedFormulaDatat_to_noface_Add_to_Temp(exfileReadFormulaData,excel_file_temp):
    print("copying richard challenge run data")
    try:
        sheetColumnFlag = ''
        sheetRowFlag = 0
        FlagFileName = exfileReadFormulaData
        if "Noface_Challenge_A_20lux_720p_Run_1" in FlagFileName:
                sheetColumnFlag = 7
                sheetRowFlag = 6
        elif "Noface_Challenge_A_20lux_720p_Run_2" in FlagFileName:
                sheetColumnFlag = 8
                sheetRowFlag = 6
        elif "Noface_Challenge_A_20lux_720p_Run_3" in FlagFileName:
                sheetColumnFlag = 9
                sheetRowFlag = 6         
        elif "Noface_Challenge_A_20lux_1080p_Run_1" in FlagFileName:
                sheetColumnFlag = 4
                sheetRowFlag = 6  
        elif "Noface_Challenge_A_20lux_1080p_Run_2" in FlagFileName:
                sheetColumnFlag = 5
                sheetRowFlag = 6 
        elif "Noface_Challenge_A_20lux_1080p_Run_3" in FlagFileName:
                sheetColumnFlag = 6
                sheetRowFlag = 6 
        elif "Noface_Challenge_CW_80lux_720p_Run_1" in FlagFileName:
                sheetColumnFlag = 7
                sheetRowFlag = 42
        elif "Noface_Challenge_CW_80lux_720p_Run_2" in FlagFileName:
                sheetColumnFlag = 8
                sheetRowFlag = 42 
        elif "Noface_Challenge_CW_80lux_720p_Run_3" in FlagFileName:
                sheetColumnFlag = 9
                sheetRowFlag = 42  
        elif "Noface_Challenge_CW_80lux_1080p_Run_1" in FlagFileName:        
                sheetColumnFlag = 4
                sheetRowFlag = 42 
        elif "Noface_Challenge_CW_80lux_1080p_Run_2" in FlagFileName: 
                sheetColumnFlag = 5
                sheetRowFlag = 42 
                print('Sheet E42') 
        elif "Noface_Challenge_CW_80lux_1080p_Run_3" in FlagFileName: 
                sheetColumnFlag = 6
                sheetRowFlag = 42 
        elif "Noface_Challenge_D65_250lux_720p_Run_1" in FlagFileName: 
                sheetColumnFlag = 7
                sheetRowFlag = 78 
        elif "Noface_Challenge_D65_250lux_720p_Run_2" in FlagFileName: 
                sheetColumnFlag = 8
                sheetRowFlag = 78
        elif "Noface_Challenge_D65_250lux_720p_Run_3" in FlagFileName: 
                sheetColumnFlag = 9
                sheetRowFlag = 78   
        elif "Noface_Challenge_D65_250lux_1080p_Run_1" in FlagFileName:      
                sheetColumnFlag = 4
                sheetRowFlag = 78 
        elif "Noface_Challenge_D65_250lux_1080p_Run_2" in FlagFileName:  
                sheetColumnFlag = 5
                sheetRowFlag = 78 
        elif "Noface_Challenge_D65_250lux_1080p_Run_3" in FlagFileName: 
                sheetColumnFlag = 6
                sheetRowFlag = 78


        # wxb1 = load_workbook(exfileReadFormulaData)
        # wxb1s = wxb1.worksheets[0]
        wxb1 = xw.Book(exfileReadFormulaData)
        wxb1s = wxb1.sheets['Sheet1']

        # list is declared
        NofaceExcelData = [] 
        for i, row in enumerate(range(7, 43)):
            NofaceExcelData.append(wxb1s.range(row, 1).value)
        wxb1.close()


        wxbtemp = load_workbook(excel_file_temp)
        wxb1s = wxbtemp.worksheets[1]
        n = 0
        for Excelvalue in NofaceExcelData:           
            wxb1s.cell(row= n+sheetRowFlag, column= sheetColumnFlag).value = Excelvalue
            n += 1
            
        wxbtemp.save(excel_file_temp)
        wxbtemp.close()


        print("Data has been copied successfully to temp")


    except Exception as e:
        print("An error occurred while reading and updaing temp excel:", e)
        error_message  = str(e)
        start_index = error_message.find("'") + 1
        end_index = error_message.rfind(".xlsx'") + len(".xlsx")
        extracted_filename = error_message[start_index:end_index]
        print(extracted_filename)
        # closing workbooks if opened with same names




